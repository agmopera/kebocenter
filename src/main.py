from datetime import datetime, timedelta
import json
import random
from io import BytesIO

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    abort,
    flash,
    send_file,
    jsonify
)

from database import (
    connect_database,
    create_tables,

    # ======================================
    # ŞUBELER
    # ======================================

    add_sube,
    get_subeler,
    get_sube,
    update_sube,
    delete_sube,

    # ======================================
    # ÜRÜNLER
    # ======================================

    add_urun,
    get_urunler,
    get_urun,
    update_urun,
    delete_urun,

    # ======================================
    # SİPARİŞLER
    # ======================================

    add_siparis,
    add_siparis_detay,
    get_siparisler,
    get_sube_siparisleri,
    get_onay_bekleyen_siparisler,
    get_siparis,
    get_siparis_detaylari,
    get_siparis_sube_id,
    get_siparis_detaylari_sevkiyat,

    get_tum_urunler,
    get_tum_subeler,
    get_siparis_miktari,

    delete_siparis,
    update_siparis,
    update_siparis_detay,
    delete_siparis_detay,

    siparis_durum_guncelle,
    siparis_gecmisi_ekle,
    get_siparis_gecmisi,

    # ======================================
    # FİNANS
    # ======================================

    get_finans_urunleri,
    get_finans_subeleri,
    fiyat_guncelle,
    limit_guncelle,
    aktif_siparis_var_mi,
    get_sube_limit,
    get_sube_sevkiyat_bilgileri,

    # ======================================
    # SİPARİŞ ÜRÜN İŞLEMLERİ
    # ======================================

    siparise_urun_ekle,
    siparis_stoklarini_aktar,
    sube_stok_azalt,
    get_sube_urun_id,
    get_sube_stoklari,
    get_stok_hareketleri,
    get_merkez_stok,
    sipariste_urun_var_mi,
    toplam_koli,
    merkez_stok_giris,
    stok_hareketi_ekle,

    # ======================================
    # SEVKİYAT
    # ======================================

    sevkiyat_programi_ekle,
    get_sevkiyat_programi,
    sevkiyat_durum_guncelle,

    # ======================================
    # GİRİŞ
    # ======================================

    login_kontrol,

    # ======================================
    # DASHBOARD
    # ======================================

    toplam_sube_sayisi,
    toplam_urun_sayisi,
    toplam_siparis_sayisi
)


# ==========================================================
# FLASK UYGULAMASI
# ==========================================================

app = Flask(
    __name__,
    template_folder="../templates",
    static_folder="../static"
)

app.secret_key = "siparis_sistemi_2026"


# ==========================================================
# DEBUG / REQUEST KONTROL
# ==========================================================

@app.before_request
def test():
    print(
        "METHOD:",
        request.method,
        "URL:",
        request.path
    )


# ==========================================================
# VERİTABANI
# ==========================================================

create_tables()


# ==========================================================
# SİPARİŞ DURUMLARI
# ==========================================================

SIPARIS_DURUMLARI = [
    "Hazırlandı",
    "Onaylandı",
    "Hazırlanıyor",
    "Sevk edildi",
    "Tamamlandı"
]


# ==========================================================
# YETKİ KONTROL
# ==========================================================

def admin_kontrol():

    if "giris" not in session:
        return redirect("/")

    if session.get("yetki") != "admin":
        abort(403)

    return None


# ==========================================================
# LOGIN
# ==========================================================

@app.route("/", methods=["GET", "POST"])
def login():

    print("LOGIN FONKSİYONU ÇALIŞTI")

    if request.method == "POST":

        kullanici_adi = request.form.get(
            "kullanici",
            ""
        ).strip()

        sifre = request.form.get(
            "sifre",
            ""
        )

        print(
            "GELEN:",
            kullanici_adi,
            sifre
        )

        # ==========================================
        # ADMIN GİRİŞİ
        # ==========================================

        if (
            kullanici_adi == "admin"
            and sifre == "admin123"
        ):

            session.clear()

            session["giris"] = True
            session["yetki"] = "admin"
            session["sube_id"] = 0
            session["sube_adi"] = "ADMIN"

            return redirect("/dashboard")

        # ==========================================
        # ŞUBE GİRİŞİ
        # ==========================================

        kullanici = login_kontrol(
            kullanici_adi,
            sifre
        )

        if kullanici:

            session.clear()

            session["giris"] = True
            session["yetki"] = "sube"
            session["sube_id"] = kullanici[0]
            session["sube_adi"] = kullanici[1]

            return redirect("/dashboard")

        return """
        <h2>Kullanıcı adı veya şifre hatalı.</h2>
        <a href="/">Tekrar Dene</a>
        """

    return render_template("login.html")


# ==========================================================
# ÇIKIŞ
# ==========================================================

@app.route("/logout")
def logout():

    session.clear()

    return redirect("/")


# ==========================================================
# DASHBOARD
# ==========================================================

@app.route("/dashboard")
def dashboard():

    if "giris" not in session:
        return redirect("/")

    return render_template(
        "dashboard.html",
        yetki=session["yetki"],
        sube=session["sube_adi"],
        siparis_sayisi=toplam_siparis_sayisi(),
        sube_sayisi=toplam_sube_sayisi(),
        urun_sayisi=toplam_urun_sayisi()
    )


# ==========================================================
# ŞUBELER
# ==========================================================

@app.route("/subeler")
def subeler():

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    liste = get_subeler()

    return render_template(
        "subeler.html",
        subeler=liste
    )


# ==========================================================
# YENİ ŞUBE
# ==========================================================

@app.route("/yeni-sube", methods=["GET", "POST"])
def yeni_sube():

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    if request.method == "POST":

        add_sube(
            request.form.get("sube_adi", ""),
            request.form.get("kullanici_adi", ""),
            request.form.get("sifre", ""),
            request.form.get("yetkili", ""),
            request.form.get("telefon", ""),
            request.form.get("eposta", ""),
            request.form.get("il", ""),
            request.form.get("ilce", ""),
            request.form.get("adres", ""),
            request.form.get("durum", ""),
            request.form.get("uretim_gunu", ""),
            request.form.get("sevkiyat_gunu", ""),
            request.form.get("sevkiyat_saati", ""),
            request.form.get("teslim_suresi", ""),
            request.form.get("firma", ""),
            request.form.get("lojistik_bedeli", 0)
        )

        return redirect("/subeler")

    return render_template("yeni_sube.html")


# ==========================================================
# ŞUBE DÜZENLE
# ==========================================================

@app.route(
    "/sube-duzenle/<int:id>",
    methods=["GET", "POST"]
)
def sube_duzenle(id):

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    if request.method == "POST":

        update_sube(
            id,
            request.form.get("sube_adi", ""),
            request.form.get("kullanici_adi", ""),
            request.form.get("sifre", ""),
            request.form.get("yetkili", ""),
            request.form.get("telefon", ""),
            request.form.get("eposta", ""),
            request.form.get("il", ""),
            request.form.get("ilce", ""),
            request.form.get("adres", ""),
            request.form.get("durum", ""),
            request.form.get("uretim_gunu", ""),
            request.form.get("sevkiyat_gunu", ""),
            request.form.get("sevkiyat_saati", ""),
            request.form.get("teslim_suresi", ""),
            request.form.get("firma", ""),
            request.form.get("lojistik_bedeli", 0)
        )

        return redirect("/subeler")

    sube = get_sube(id)

    if not sube:
        abort(404)

    return render_template(
        "sube_duzenle.html",
        sube=sube
    )


# ==========================================================
# ŞUBE SİL
# ==========================================================

@app.route("/sube-sil/<int:id>")
def sube_sil(id):

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    delete_sube(id)

    return redirect("/subeler")


# ==========================================================
# ÜRÜNLER
# ==========================================================

@app.route("/urunler")
def urunler():

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    liste = get_urunler()

    return render_template(
        "urunler.html",
        urunler=liste
    )


# ==========================================================
# FİNANS
# ==========================================================

@app.route("/finans", methods=["GET", "POST"])
def finans():

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    if request.method == "POST":

        islem = request.form.get("islem")

        # ==========================================
        # FİYATLARI KAYDET
        # ==========================================

        if islem == "fiyatlar":

            urunler = get_finans_urunleri()

            for urun in urunler:

                fiyat = request.form.get(
                    f"fiyat_{urun[0]}",
                    "0"
                )

                try:
                    fiyat = float(fiyat)
                except (ValueError, TypeError):
                    fiyat = 0

                fiyat_guncelle(
                    urun[0],
                    fiyat
                )

        # ==========================================
        # LİMİTLERİ KAYDET
        # ==========================================

        elif islem == "limitler":

            subeler = get_finans_subeleri()

            for sube in subeler:

                limit = request.form.get(
                    f"limit_{sube[0]}",
                    "0"
                )

                try:
                    limit = float(limit)
                except (ValueError, TypeError):
                    limit = 0

                limit_guncelle(
                    sube[0],
                    limit
                )

        return redirect("/finans")

    urunler = get_finans_urunleri()
    subeler = get_finans_subeleri()

    return render_template(
        "finans.html",
        urunler=urunler,
        subeler=subeler
    )


# ==========================================================
# RAPORLAR
# ==========================================================

@app.route("/raporlar")
def raporlar():

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    return render_template("raporlar.html")


# ==========================================================
# AI
# ==========================================================

@app.route("/ai")
def ai():

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    return render_template("ai.html")


# ==========================================================
# MERKEZ STOK GİRİŞİ
# ==========================================================

@app.route(
    "/stok-giris",
    methods=["POST"]
)
def stok_giris():

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    try:

        urun_id = int(
            request.form.get("urun_id")
        )

        miktar = float(
            request.form.get("miktar", 0)
        )

    except (ValueError, TypeError):

        return redirect("/stoklar")

    if miktar <= 0:
        return redirect("/stoklar")

    merkez_stok_giris(
        urun_id,
        miktar
    )

    stok_hareketi_ekle(
        urun_id,
        miktar,
        "Giriş",
        "Tedarikçi",
        "Merkez Depo",
        None,
        "Merkez stok girişi"
    )

    return redirect("/stoklar")


# ======================================
# ÜRÜN TEDARİK ŞEKLİ YARDIMCILARI
# ======================================

def urun_tedarik_sekli_guncelle(urun_id, tedarik_sekli):
    """Mevcut database.py fonksiyonlarını bozmadan tedarik şeklini kaydeder."""
    conn = connect_database()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE urunler
        SET tedarik_sekli=?
        WHERE id=?
    """, (tedarik_sekli, urun_id))
    conn.commit()
    conn.close()


def urun_tedarik_sekli_getir(urun_id):
    conn = connect_database()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT COALESCE(tedarik_sekli, 'Merkez')
        FROM urunler
        WHERE id=?
    """, (urun_id,))
    sonuc = cursor.fetchone()
    conn.close()
    return sonuc[0] if sonuc else 'Merkez'


def get_tum_urunler_siparis():
    """Yeni sipariş için mevcut ürün alanlarını korur ve tedarik şeklini sona ekler."""
    conn = connect_database()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT
            id,
            urun_adi,
            kategori,
            birim,
            palet_kapasitesi,
            fiyat,
            urun_tipi,
            COALESCE(tedarik_sekli, 'Merkez')
        FROM urunler
        WHERE durum='Aktif'
        ORDER BY urun_adi
    """)
    sonuc = cursor.fetchall()
    conn.close()
    return sonuc


# ======================================
# YENİ ÜRÜN EKLE
# ======================================

@app.route("/yeni-urun", methods=["GET", "POST"])
def yeni_urun():

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    if request.method == "POST":

        urun_kodu = request.form.get("urun_kodu", "").strip()
        urun_adi = request.form.get("urun_adi", "").strip()
        kategori = request.form.get("kategori", "").strip()
        birim = request.form.get("birim", "").strip()
        durum = request.form.get("durum", "Aktif").strip()

        palet_kapasitesi = request.form.get(
            "palet_kapasitesi",
            0
        )

        urun_tipi = request.form.get(
            "urun_tipi",
            "Donuk"
        ).strip()

        koli_agirligi = request.form.get(
            "koli_agirligi",
            0
        )

        # ======================================
        # TEDARİK ŞEKLİ
        # ======================================

        tedarik_sekli = request.form.get(
            "tedarik_sekli",
            "Merkez"
        ).strip()

        # ======================================
        # ÜRÜNÜ VERİTABANINA EKLE
        # ======================================

        add_urun(
            urun_kodu,
            urun_adi,
            kategori,
            birim,
            durum,
            palet_kapasitesi,
            urun_tipi,
            koli_agirligi,
            tedarik_sekli
        )

        return redirect("/urunler")

    return render_template(
        "yeni_urun.html"
    )
    

# ==========================================================
# ÜRÜN DÜZENLE
# ==========================================================
@app.route(
    "/urun-duzenle/<int:id>",
    methods=["GET", "POST"]
)
def urun_duzenle(id):

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    if request.method == "POST":

        update_urun(
            id,
            request.form.get("urun_kodu", "").strip(),
            request.form.get("urun_adi", "").strip(),
            request.form.get("kategori", "").strip(),
            request.form.get("birim", "").strip(),
            request.form.get("palet_kapasitesi", 0),
            request.form.get("urun_tipi", "").strip(),
            request.form.get("koli_agirligi", 0),
            request.form.get("durum", "").strip(),
            request.form.get("tedarik_sekli", "Merkez").strip()
        )

        return redirect("/urunler")

    urun = get_urun(id)

    if not urun:
        abort(404)

    return render_template(
        "urun_duzenle.html",
        urun=urun
    )

# ==========================================================
# ÜRÜN SİL
# ==========================================================

@app.route("/urun-sil/<int:id>")
def urun_sil(id):

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    delete_urun(id)

    return redirect("/urunler")


# ==========================================================
# SİPARİŞLER
# ==========================================================

@app.route("/siparisler")
def siparisler():

    if "giris" not in session:
        return redirect("/")

    if session["yetki"] == "admin":

        liste = get_siparisler()

    else:

        liste = get_sube_siparisleri(
            session["sube_id"]
        )

    return render_template(
        "siparisler.html",
        siparisler=liste,
        yetki=session["yetki"]
    )


# ==========================================================
# SİPARİŞ DETAY
# ==========================================================

@app.route("/siparis-detay/<int:id>")
def siparis_detay(id):

    if "giris" not in session:
        return redirect("/")

    siparis = get_siparis(id)

    if not siparis:
        abort(404)

    # ==========================================
    # ŞUBE YETKİ KONTROLÜ
    # ==========================================

    if session["yetki"] == "sube":

        if siparis[2] != session["sube_adi"]:
            abort(403)

    detaylar = get_siparis_detaylari(id)

    gecmis = get_siparis_gecmisi(id)

    return render_template(
        "siparis_detay.html",
        siparis=siparis,
        detaylar=detaylar,
        gecmis=gecmis
    )


# ==========================================================
# SİPARİŞ SİL
# ==========================================================

@app.route("/siparis-sil/<int:id>")
def siparis_sil(id):

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    delete_siparis(id)

    return redirect("/siparisler")


# ==========================================================
# SİPARİŞ DURUMU
# ==========================================================

@app.route(
    "/siparis-durum/<int:id>",
    methods=["POST"]
)
def siparis_durum(id):

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    siparis = get_siparis(id)

    if not siparis:
        abort(404)

    eski_durum = siparis[4]

    yeni_durum = request.form.get("durum")

    if not yeni_durum:
        return redirect("/siparisler")

    # ==========================================
    # DURUM DEĞİŞMEDİYSE KAYIT OLUŞTURMA
    # ==========================================

    if eski_durum == yeni_durum:
        return redirect("/siparisler")

    siparis_durum_guncelle(
        id,
        yeni_durum
    )

    siparis_gecmisi_ekle(
        siparis_id=id,
        tarih=datetime.now().strftime(
            "%d.%m.%Y %H:%M"
        ),
        kullanici=session.get(
            "sube_adi",
            "Admin"
        ),
        eski_durum=eski_durum,
        yeni_durum=yeni_durum
    )

    return redirect("/siparisler")


# ==========================================================
# SEVKİYAT TARİH / SAAT HESABI
# ==========================================================

def sevkiyat_zamani_hesapla(
    siparis_tarihi,
    sevkiyat_gunu,
    sevkiyat_saati,
    teslim_suresi
):

    try:

        # ==========================================
        # SİPARİŞ TARİHİ
        # ==========================================

        tarih = datetime.strptime(
            str(siparis_tarihi),
            "%Y-%m-%d"
        )

        # ==========================================
        # SEVKİYAT GÜNLERİ
        # ==========================================

        gunler = {
            "Pazartesi": 0,
            "Salı": 1,
            "Çarşamba": 2,
            "Perşembe": 3,
            "Cuma": 4,
            "Cumartesi": 5,
            "Pazar": 6
        }

        hedef_gun = gunler.get(
            sevkiyat_gunu
        )

        if hedef_gun is None:
            return None, None, None

        # ==========================================
        # BİR SONRAKİ SEVKİYAT GÜNÜ
        # ==========================================

        mevcut_gun = tarih.weekday()

        fark = (
            hedef_gun - mevcut_gun
        ) % 7

        sevkiyat_tarihi = (
            tarih +
            timedelta(days=fark)
        )

        # ==========================================
        # SEVKİYAT SAATİ
        # ==========================================

        if not sevkiyat_saati:
            sevkiyat_saati = "18:00"

        saat = datetime.strptime(
            str(sevkiyat_saati),
            "%H:%M"
        ).time()

        atb_cikis = datetime.combine(
            sevkiyat_tarihi.date(),
            saat
        )

        # ==========================================
        # TESLİM SÜRESİ
        # ==========================================

        try:

            teslim_suresi = float(
                teslim_suresi or 0
            )

        except (ValueError, TypeError):

            teslim_suresi = 0

        # ==========================================
        # ŞUBE TESLİM
        # ==========================================

        sube_teslim = (
            atb_cikis +
            timedelta(hours=teslim_suresi)
        )

        return (
            sevkiyat_tarihi,
            atb_cikis,
            sube_teslim
        )

    except Exception as e:

        print(
            "Sevkiyat zamanı hesaplama hatası:",
            e
        )

        return None, None, None


# ==========================================================
# ADMIN SİPARİŞ ONAY
# ==========================================================

@app.route(
    "/siparis-onayla/<int:id>",
    methods=["GET", "POST"]
)
def siparis_onayla(id):

    # ==========================================
    # ADMIN KONTROLÜ
    # ==========================================

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    # ==========================================
    # SİPARİŞİ BUL
    # ==========================================

    siparis = get_siparis(id)

    if not siparis:
        abort(404)

    # ==========================================
    # SADECE HAZIRLANDI DURUMUNDA ONAY
    # ==========================================

    if siparis[4] != "Hazırlandı":
        return redirect("/siparisler")

    # ==========================================
    # ŞUBE ID
    # ==========================================

    sube_id = get_siparis_sube_id(id)

    if not sube_id:
        return redirect("/siparisler")

    # ==========================================
    # SİPARİŞ TARİHİ
    # ==========================================

    siparis_tarihi = siparis[3]

    # ==========================================
    # ŞUBE SEVKİYAT BİLGİLERİ
    # ==========================================

    sube_bilgileri = get_sube_sevkiyat_bilgileri(
        sube_id
    )

    if not sube_bilgileri:
        return redirect("/siparisler")

    sevkiyat_gunu = sube_bilgileri[0]
    sevkiyat_saati = sube_bilgileri[1]
    teslim_suresi = sube_bilgileri[2]
    firma = sube_bilgileri[3]

    try:

        lojistik_bedeli = float(
            sube_bilgileri[4] or 0
        )

    except (ValueError, TypeError):

        lojistik_bedeli = 0

    # ==========================================
    # SİPARİŞ DETAYLARI
    # ==========================================

    detaylar = get_siparis_detaylari_sevkiyat(
        id
    )

    # ==========================================
    # PALET TOPLAMLARI
    # ==========================================

    donuk = 0
    soguk = 0

    donuk_esdeger = 0
    soguk_esdeger = 0

    # ==========================================
    # ÜRÜNLERİ TEK TEK İNCELE
    # ==========================================

    for detay in detaylar:

        # --------------------------------------
        # detay:
        #
        # 0 = urun_id
        # 1 = urun_adi
        # 2 = urun_tipi
        # 3 = palet_kapasitesi
        # 4 = miktar
        # --------------------------------------

        urun_tipi = detay[2]

        try:

            palet_kapasitesi = float(
                detay[3] or 0
            )

        except (ValueError, TypeError):

            palet_kapasitesi = 0

        try:

            miktar = float(
                detay[4] or 0
            )

        except (ValueError, TypeError):

            miktar = 0

        # ==========================================
        # TEDARİK ŞEKLİ
        # ==========================================
        # Ürün "Tedarikçi" seçilmişse bu ürün
        # doğrudan tedarikçiden gelir ve merkezin
        # paletine dahil edilmez.
        #
        # Bu bilgi sevkiyat ekranında gösterilmez.
        # Sadece palet doluluk hesabında kullanılır.
        # ==========================================

        tedarik_sekli = urun_tedarik_sekli_getir(
            detay[0]
        )

        if str(tedarik_sekli).strip().lower() == "tedarikçi":
            continue

        # ==========================================
        # PALET KAPASİTESİ YOKSA GEÇ
        # ==========================================

        if palet_kapasitesi <= 0:
            continue

        # ==========================================
        # PALET EŞDEĞERİ
        # ==========================================

        palet_esdegeri = (
            miktar / palet_kapasitesi
        )

        # ==========================================
        # EXCEL YUVARLAMA KURALI
        #
        # 0,00 - 0,11 -> aşağı
        # 0,12 ve üstü -> yukarı
        # ==========================================

        tam_kisim = int(
            palet_esdegeri
        )

        ondalik_kisim = (
            palet_esdegeri
            - tam_kisim
        )

        if ondalik_kisim < 0.12:

            fiziksel_palet = tam_kisim

        else:

            fiziksel_palet = tam_kisim + 1

        # ==========================================
        # MİKTAR VARSA EN AZ 1 PALET
        # ==========================================

        if (
            miktar > 0
            and fiziksel_palet == 0
        ):

            fiziksel_palet = 1

        # ==========================================
        # DONUK
        # ==========================================

        if urun_tipi == "Donuk":

            donuk += fiziksel_palet

            donuk_esdeger += palet_esdegeri

        # ==========================================
        # SOĞUK
        # ==========================================

        elif urun_tipi == "Soğuk":

            soguk += fiziksel_palet

            soguk_esdeger += palet_esdegeri

    # ==========================================
    # DOLULUK HESAPLARI
    # ==========================================

    donuk_doluluk = 0
    soguk_doluluk = 0

    if donuk > 0:

        donuk_doluluk = (
            donuk_esdeger / donuk
        ) * 100

    if soguk > 0:

        soguk_doluluk = (
            soguk_esdeger / soguk
        ) * 100

    # ==========================================
    # LOJİSTİK MALİYETİ
    # ==========================================

    toplam_palet = (
        donuk + soguk
    )

    toplam_maliyet = (
        toplam_palet
        * lojistik_bedeli
    )

    # ==========================================
    # KAYIP MALİYETİ
    # ==========================================

    donuk_kayip_maliyeti = (
        donuk
        * (
            1
            - (donuk_doluluk / 100)
        )
        * lojistik_bedeli
    )

    soguk_kayip_maliyeti = (
        soguk
        * (
            1
            - (soguk_doluluk / 100)
        )
        * lojistik_bedeli
    )

    kayip_maliyeti = (
        donuk_kayip_maliyeti
        + soguk_kayip_maliyeti
    )

    # ==========================================
    # SEVKİYAT ZAMANI
    # ==========================================

    (
        sevkiyat_tarihi,
        atb_cikis,
        sube_teslim
    ) = sevkiyat_zamani_hesapla(
        siparis_tarihi,
        sevkiyat_gunu,
        sevkiyat_saati,
        teslim_suresi
    )

    if not sevkiyat_tarihi:
        return redirect("/siparisler")

    # ==========================================
    # HAFTA
    # ==========================================

    hafta = (
        sevkiyat_tarihi.isocalendar().week
    )

    # ==========================================
    # SEVKİYAT PROGRAMINA KAYIT
    # ==========================================
    #
    # YENİ:
    #
    # ==========================================

    sevkiyat_programi_ekle(
        hafta=hafta,
        tarih=sevkiyat_tarihi,
        firma=firma,
        sube_id=sube_id,
        donuk=donuk,
        soguk=soguk,
        donuk_doluluk=donuk_doluluk,
        soguk_doluluk=soguk_doluluk,
        atb_cikis=atb_cikis.strftime(
            "%d.%m.%Y %H:%M"
        ),
        sube_teslim=sube_teslim.strftime(
            "%d.%m.%Y %H:%M"
        ),
        palet_fiyat=lojistik_bedeli,
        toplam_maliyet=toplam_maliyet,
        kayip_maliyeti=kayip_maliyeti,
        durum="AKTIF"
    )

    # ==========================================
    # STOKLARI AKTAR
    # ==========================================

    siparis_stoklarini_aktar(id)

    # ==========================================
    # SİPARİŞ DURUMU
    # ==========================================

    siparis_durum_guncelle(
        id,
        "Onaylandı"
    )

    # ==========================================
    # SİPARİŞ GEÇMİŞİ
    # ==========================================

    siparis_gecmisi_ekle(
        siparis_id=id,
        tarih=datetime.now().strftime(
            "%d.%m.%Y %H:%M"
        ),
        kullanici=session.get(
            "sube_adi",
            "Admin"
        ),
        eski_durum="Hazırlandı",
        yeni_durum="Onaylandı"
    )

    return redirect("/siparisler")


# ==========================================================
# SİPARİŞ DÜZENLE
# ==========================================================

@app.route(
    "/siparis-duzenle/<int:id>",
    methods=["GET", "POST"]
)
def siparis_duzenle(id):

    if "giris" not in session:
        return redirect("/")

    siparis = get_siparis(id)

    if not siparis:
        return "Sipariş bulunamadı"

    durum = siparis[4]

    # ==========================================
    # ŞUBE KONTROLÜ
    # ==========================================

    if session.get("yetki") == "sube":

        if siparis[2] != session.get("sube_adi"):
            abort(403)

        if durum != "Hazırlandı":

            return """
            <h3>
            Bu sipariş artık düzenlenemez.
            </h3>

            <a href="/siparisler">
            Siparişlere Dön
            </a>
            """

    # ==========================================
    # POST
    # ==========================================

    if request.method == "POST":

        detaylar = get_siparis_detaylari(id)

        # ==========================================
        # MEVCUT ÜRÜNLERİ GÜNCELLE
        # ==========================================

        for detay in detaylar:

            yeni_miktar = request.form.get(
                f"miktar_{detay[0]}"
            )

            if yeni_miktar is not None:

                update_siparis_detay(
                    detay[0],
                    yeni_miktar
                )

        # ==========================================
        # YENİ ÜRÜN
        # ==========================================

        yeni_urun = request.form.get(
            "yeni_urun"
        )

        yeni_miktar = request.form.get(
            "yeni_miktar"
        )

        if (
            yeni_urun
            and yeni_miktar
        ):

            if not sipariste_urun_var_mi(
                id,
                yeni_urun
            ):

                siparise_urun_ekle(
                    id,
                    yeni_urun,
                    yeni_miktar
                )

        return redirect(
            f"/siparis-duzenle/{id}"
        )

    # ==========================================
    # GET
    # ==========================================

    detaylar = get_siparis_detaylari(id)

    urunler = get_urunler()

    toplam = toplam_koli(id)

    return render_template(
        "siparis_duzenle.html",
        siparis=siparis,
        detaylar=detaylar,
        urunler=urunler,
        toplam=toplam
    )


# ==========================================================
# SİPARİŞ DETAY SİL
# ==========================================================

@app.route(
    "/siparis-detay-sil/<int:detay_id>/<int:siparis_id>"
)
def siparis_detay_sil(
    detay_id,
    siparis_id
):

    # ==========================================
    # GİRİŞ KONTROLÜ
    # ==========================================

    if "giris" not in session:
        return redirect("/")

    # ==========================================
    # ŞUBE KONTROLÜ
    # ==========================================

    if session.get("yetki") == "sube":

        siparis = get_siparis(siparis_id)

        if not siparis:
            abort(404)

        if siparis[2] != session.get("sube_adi"):
            abort(403)

        if siparis[4] != "Hazırlandı":
            abort(403)

    else:

        sonuc = admin_kontrol()

        if sonuc:
            return sonuc

    delete_siparis_detay(
        detay_id
    )

    return redirect(
        f"/siparis-duzenle/{siparis_id}"
    )


# ==========================================================
# YENİ SİPARİŞ
# ==========================================================

@app.route(
    "/yeni-siparis",
    methods=["GET", "POST"]
)
def yeni_siparis():

    if "giris" not in session:
        return redirect("/")

    # ==========================================
    # POST
    # ==========================================

    if request.method == "POST":

        # ==========================================
        # ADMIN İSTEDİĞİ ŞUBEYİ SEÇEBİLİR
        # ==========================================

        if session["yetki"] == "admin":

            try:

                sube_id = int(
                    request.form["sube_id"]
                )

            except (ValueError, TypeError, KeyError):

                return redirect("/yeni-siparis")

        # ==========================================
        # ŞUBE SADECE KENDİ ADINA SİPARİŞ VERİR
        # ==========================================

        else:

            sube_id = session["sube_id"]

            # ======================================
            # AKTİF SİPARİŞ KONTROLÜ
            # ======================================

            if aktif_siparis_var_mi(
                sube_id
            ):

                return """
                <script>
                    alert(
                        "Bu hafta zaten aktif bir siparişiniz bulunmaktadır."
                    );
                    window.location="/siparisler";
                </script>
                """

        # ==========================================
        # TARİH
        # ==========================================

        tarih = datetime.now().strftime(
            "%Y-%m-%d"
        )

        # ==========================================
        # ÜRÜNLER
        # ==========================================

        try:

            urunler_json = request.form[
                "urunler_json"
            ]

            urun_listesi = json.loads(
                urunler_json
            )

        except (
            KeyError,
            json.JSONDecodeError,
            TypeError
        ):

            return redirect(
                "/yeni-siparis"
            )

        if not isinstance(
            urun_listesi,
            list
        ):

            return redirect(
                "/yeni-siparis"
            )

        if len(urun_listesi) == 0:

            return redirect(
                "/yeni-siparis"
            )

        # ==========================================
        # SİPARİŞ NUMARASI
        # ==========================================

        siparis_no = (
            "SP-"
            + datetime.now().strftime("%Y%m%d")
            + "-"
            + str(
                random.randint(
                    1000,
                    9999
                )
            )
        )

        # ==========================================
        # SİPARİŞİ OLUŞTUR
        # ==========================================

        siparis_id = add_siparis(
            siparis_no,
            sube_id,
            tarih,
            "Hazırlandı"
        )

        # ==========================================
        # SİPARİŞ DETAYLARI
        # ==========================================

        for urun in urun_listesi:

            try:

                urun_id = int(
                    urun["id"]
                )

                miktar = float(
                    urun["miktar"]
                )

            except (
                KeyError,
                ValueError,
                TypeError
            ):

                continue

            if miktar <= 0:
                continue

            add_siparis_detay(
                siparis_id,
                urun_id,
                miktar
            )

        return redirect("/siparisler")

    # ==========================================
    # GET
    # ==========================================

    if session["yetki"] == "admin":

        subeler = get_subeler()

        if not subeler:

            return """
            <h3>
            Sistemde kayıtlı şube bulunmamaktadır.
            </h3>

            <a href="/subeler">
            Şubelere Dön
            </a>
            """

    else:

        # ==========================================
        # AKTİF SİPARİŞ KONTROLÜ
        # ==========================================

        if aktif_siparis_var_mi(
            session["sube_id"]
        ):

            return """
            <script>
                alert(
                    "Bu hafta zaten aktif bir siparişiniz bulunmaktadır."
                );
                window.location="/siparisler";
            </script>
            """

        subeler = [
            (
                session["sube_id"],
                session["sube_adi"]
            )
        ]

    # ==========================================
    # ÜRÜNLER
    # ==========================================

    urunler = get_tum_urunler_siparis()

    # ==========================================
    # ŞUBE FİNANS LİMİTİ
    # ==========================================

    if session["yetki"] == "admin":

        secili_sube = subeler[0][0]

    else:

        secili_sube = session["sube_id"]

    limit = get_sube_limit(
        secili_sube
    )

    # ==========================================
    # SAYFA
    # ==========================================

    return render_template(
        "yeni_siparis.html",
        subeler=subeler,
        urunler=urunler,
        limit=limit
    )


# ==========================================================
# ŞUBE LİMİT GETİRME - AJAX
# ==========================================================

@app.route(
    "/get-sube-limit/<int:sube_id>"
)
def get_sube_limit_ajax(sube_id):

    if "giris" not in session:

        return jsonify({
            "limit": 0
        })

    # ==========================================
    # ŞUBE KULLANICISI SADECE KENDİ LİMİTİNİ
    # GÖREBİLİR
    # ==========================================

    if (
        session.get("yetki") == "sube"
        and sube_id != session.get("sube_id")
    ):

        return jsonify({
            "limit": 0
        })

    limit = get_sube_limit(
        sube_id
    )

    return jsonify({
        "limit": limit or 0
    })


# ==========================================================
# STOKLAR
# ==========================================================

@app.route("/stoklar")
def stoklar():

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    # ==========================================
    # FİLTRELER
    # ==========================================

    baslangic = request.args.get(
        "baslangic",
        ""
    )

    bitis = request.args.get(
        "bitis",
        ""
    )

    sube_id = request.args.get(
        "sube_id",
        ""
    )

    urun_id = request.args.get(
        "urun_id",
        ""
    )

    hareket_tipi = request.args.get(
        "hareket_tipi",
        ""
    )

    arama = request.args.get(
        "arama",
        ""
    )

    # ==========================================
    # BOŞ DEĞERLERİ NONE YAP
    # ==========================================

    try:

        sube_id = (
            int(sube_id)
            if sube_id
            else None
        )

    except (ValueError, TypeError):

        sube_id = None

    try:

        urun_id = (
            int(urun_id)
            if urun_id
            else None
        )

    except (ValueError, TypeError):

        urun_id = None

    # ==========================================
    # MERKEZ STOKLARI
    # ==========================================

    merkez_stok = get_merkez_stok()

    # ==========================================
    # ŞUBE STOKLARI
    # ==========================================

    sube_stoklari = get_sube_stoklari(
        sube_id
    )

    # ==========================================
    # STOK HAREKETLERİ
    # ==========================================

    stok_hareketleri = get_stok_hareketleri(
        baslangic or None,
        bitis or None,
        sube_id,
        urun_id,
        hareket_tipi or None,
        arama.strip() or None
    )

    # ==========================================
    # FİLTRE SEÇENEKLERİ
    # ==========================================

    filtre_subeler = get_tum_subeler()

    filtre_urunler = get_tum_urunler()

    return render_template(
        "stoklar.html",
        merkez_stok=merkez_stok,
        sube_stoklari=sube_stoklari,
        stok_hareketleri=stok_hareketleri,
        filtre_subeler=filtre_subeler,
        filtre_urunler=filtre_urunler,
        baslangic=baslangic,
        bitis=bitis,
        secili_sube=sube_id,
        secili_urun=urun_id,
        secili_hareket=hareket_tipi,
        arama=arama
    )


# ==========================================================
# MERKEZ STOKLARINI EXCEL'E AKTAR
# ==========================================================

@app.route("/stoklar/merkez-excel")
def stoklar_merkez_excel():

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    # ==========================================
    # MERKEZ STOKLARINI AL
    # ==========================================

    merkez_stok = get_merkez_stok()

    # ==========================================
    # EXCEL OLUŞTUR
    # ==========================================

    wb = Workbook()

    ws = wb.active

    ws.title = "Merkez Stokları"

    # ==========================================
    # BAŞLIK
    # ==========================================

    basliklar = [
        "Ürün ID",
        "Ürün Adı",
        "Birim",
        "Merkez Stok"
    ]

    ws.append(basliklar)

    # ==========================================
    # BAŞLIK FORMAT
    # ==========================================

    for cell in ws[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    # ==========================================
    # VERİLER
    # ==========================================

    for stok in merkez_stok:

        urun_id = stok[0]
        urun_adi = stok[1]
        birim = stok[2]
        miktar = stok[3]

        ws.append([
            urun_id,
            urun_adi,
            birim,
            miktar
        ])

    # ==========================================
    # SÜTUN GENİŞLİKLERİ
    # ==========================================

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 18

    # ==========================================
    # DOSYAYI BELLEKTE OLUŞTUR
    # ==========================================

    dosya = BytesIO()

    wb.save(dosya)

    dosya.seek(0)

    return send_file(
        dosya,
        as_attachment=True,
        download_name="Merkez_Stoklari.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


# ==========================================================
# ŞUBE STOKLARINI EXCEL'E AKTAR
# ==========================================================

@app.route("/stoklar/sube-excel")
def stoklar_sube_excel():

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    # ==========================================
    # ŞUBE FİLTRESİNİ AL
    # ==========================================

    sube_id = request.args.get(
        "sube_id",
        ""
    )

    try:

        sube_id = (
            int(sube_id)
            if sube_id
            else None
        )

    except (ValueError, TypeError):

        sube_id = None

    # ==========================================
    # FİLTRELİ ŞUBE STOKLARINI AL
    # ==========================================

    sube_stoklari = get_sube_stoklari(
        sube_id
    )

    # ==========================================
    # EXCEL OLUŞTUR
    # ==========================================

    wb = Workbook()

    ws = wb.active

    ws.title = "Şube Stokları"

    # ==========================================
    # BAŞLIKLAR
    # ==========================================

    basliklar = [
        "Şube ID",
        "Şube Adı",
        "Ürün ID",
        "Ürün Adı",
        "Birim",
        "Stok Miktarı"
    ]

    ws.append(basliklar)

    # ==========================================
    # BAŞLIK FORMAT
    # ==========================================

    for cell in ws[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    # ==========================================
    # VERİLER
    # ==========================================

    for stok in sube_stoklari:

        sube_id_excel = stok[0]
        sube_adi = stok[1]
        urun_id = stok[2]
        urun_adi = stok[3]
        birim = stok[4]
        miktar = stok[5]

        ws.append([
            sube_id_excel,
            sube_adi,
            urun_id,
            urun_adi,
            birim,
            miktar
        ])

    # ==========================================
    # SÜTUN GENİŞLİKLERİ
    # ==========================================

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 18

    # ==========================================
    # DOSYA ADI
    # ==========================================

    if sube_id:

        if sube_stoklari:

            sube_adi = sube_stoklari[0][1]

            dosya_adi = (
                f"{sube_adi}_Stoklari.xlsx"
            )

        else:

            dosya_adi = "Sube_Stoklari.xlsx"

    else:

        dosya_adi = "Tum_Sube_Stoklari.xlsx"

    # ==========================================
    # DOSYAYI BELLEKTE OLUŞTUR
    # ==========================================

    dosya = BytesIO()

    wb.save(dosya)

    dosya.seek(0)

    return send_file(
        dosya,
        as_attachment=True,
        download_name=dosya_adi,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


# ==========================================================
# SEVKİYAT PROGRAMI
# ==========================================================

@app.route("/sevkiyat")
def sevkiyat():

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    # ==========================================
    # FİLTRELER
    # ==========================================

    baslangic = request.args.get(
        "baslangic",
        ""
    )

    bitis = request.args.get(
        "bitis",
        ""
    )

    sube_id = request.args.get(
        "sube_id",
        ""
    )

    firma = request.args.get(
        "firma",
        ""
    )

    # ==========================================
    # ŞUBE ID
    # ==========================================

    try:

        sube_id = (
            int(sube_id)
            if sube_id
            else None
        )

    except (ValueError, TypeError):

        sube_id = None

    # ==========================================
    # SEVKİYAT PROGRAMI
    # ==========================================

    sevkiyatlar = get_sevkiyat_programi(
        baslangic=baslangic or None,
        bitis=bitis or None,
        sube_id=sube_id,
        firma=firma or None
    )

    # ==========================================
    # FİLTRE SEÇENEKLERİ
    # ==========================================

    filtre_subeler = get_tum_subeler()

    # ==========================================
    # SAYFA
    # ==========================================

    return render_template(
        "sevkiyat.html",
        sevkiyatlar=sevkiyatlar,
        filtre_subeler=filtre_subeler,
        baslangic=baslangic,
        bitis=bitis,
        secili_sube=sube_id,
        secili_firma=firma
    )


# ==========================================================
# SEVKİYAT PROGRAMINI EXCEL'E AKTAR
# ==========================================================

@app.route("/sevkiyat-excel")
def sevkiyat_excel():

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    # ==========================================
    # FİLTRELER
    # ==========================================

    baslangic = (
        request.args.get("baslangic")
        or None
    )

    bitis = (
        request.args.get("bitis")
        or None
    )

    sube_id = (
        request.args.get("sube_id")
        or None
    )

    firma = (
        request.args.get("firma")
        or None
    )

    # ==========================================
    # SEVKİYAT VERİLERİ
    # ==========================================

    sevkiyatlar = get_sevkiyat_programi(
        baslangic=baslangic,
        bitis=bitis,
        sube_id=sube_id,
        firma=firma
    )

    # ==========================================
    # EXCEL
    # ==========================================

    wb = Workbook()

    ws = wb.active

    ws.title = "Sevkiyat Programı"

    basliklar = [
        "ID",
        "HAFTA",
        "TARİH",
        "FİRMA",
        "ŞUBE ID",
        "ŞUBE",
        "DONUK PALET",
        "SOĞUK PALET",
        "DONUK DOLULUK",
        "SOĞUK DOLULUK",
        "ATB ÇIKIŞ",
        "ŞUBE TESLİM",
        "PALET FİYATI",
        "TOPLAM MALİYET",
        "KAYIP MALİYETİ",
        "DURUM"
    ]

    ws.append(basliklar)

    # ==========================================
    # BAŞLIK STİLİ
    # ==========================================

    for cell in ws[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # ==========================================
    # VERİLER
    # ==========================================

    for sevkiyat in sevkiyatlar:

        ws.append([
            sevkiyat[0],   # ID
            sevkiyat[1],   # HAFTA
            sevkiyat[2],   # TARİH
            sevkiyat[3],   # FİRMA
            sevkiyat[4],   # ŞUBE ID
            sevkiyat[5],   # ŞUBE
            sevkiyat[6],   # DONUK
            sevkiyat[7],   # SOĞUK

            # DOLULUK
            (sevkiyat[8] or 0) / 100,
            (sevkiyat[9] or 0) / 100,

            sevkiyat[10],  # ATB ÇIKIŞ
            sevkiyat[11],  # ŞUBE TESLİM
            sevkiyat[12],  # PALET FİYATI
            sevkiyat[13],  # TOPLAM MALİYET
            sevkiyat[14],  # KAYIP MALİYETİ
            sevkiyat[15]   # DURUM
        ])

    # ==========================================
    # YÜZDE FORMATLARI
    # ==========================================

    for row in range(
        2,
        ws.max_row + 1
    ):

        ws.cell(
            row=row,
            column=10
        ).number_format = "0.00%"

        ws.cell(
            row=row,
            column=11
        ).number_format = "0.00%"

    # ==========================================
    # PARA FORMATLARI
    # ==========================================

    for row in range(
        2,
        ws.max_row + 1
    ):

        for column in [
            14,
            15,
            16
        ]:

            ws.cell(
                row=row,
                column=column
            ).number_format = (
                '#,##0.00 "₺"'
            )

    # ==========================================
    # SÜTUN GENİŞLİKLERİ
    # ==========================================

    genislikler = {
        1: 8,
        2: 10,
        3: 14,
        4: 15,
        5: 10,
        6: 28,
        7: 15,
        8: 15,
        9: 18,
        10: 18,
        11: 18,
        12: 18,
        13: 15,
        14: 18,
        15: 18,
        16: 15
    }

    for column, width in genislikler.items():

        ws.column_dimensions[
            get_column_letter(column)
        ].width = width

    # ==========================================
    # FİLTRE
    # ==========================================

    ws.auto_filter.ref = ws.dimensions

    ws.freeze_panes = "A2"

    # ==========================================
    # DOSYA
    # ==========================================

    dosya = BytesIO()

    wb.save(dosya)

    dosya.seek(0)

    return send_file(
        dosya,
        as_attachment=True,
        download_name="sevkiyat_programi.xlsx",
        mimetype=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


# ==========================================================
# SEVKİYAT DURUMU DEĞİŞTİR
# ==========================================================

@app.route(
    "/sevkiyat-durum/<int:id>",
    methods=["POST"]
)
def sevkiyat_durum(id):

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    durum = request.form.get(
        "durum"
    )

    # ==========================================
    # SADECE AKTİF / PASİF
    # ==========================================

    if durum not in [
        "AKTIF",
        "PASIF"
    ]:

        return redirect(
            request.referrer
            or "/sevkiyat"
        )

    sevkiyat_durum_guncelle(
        id,
        durum
    )

    return redirect(
        request.referrer
        or "/sevkiyat"
    )


# ==========================================================
# ŞUBE SATIŞ EXCEL AKTARIMI
# ==========================================================

@app.route(
    "/sube-satis-aktar",
    methods=["POST"]
)
def sube_satis_aktar():

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    dosya = request.files.get(
        "dosya"
    )

    # ==========================================
    # DOSYA KONTROLÜ
    # ==========================================

    if (
        not dosya
        or dosya.filename == ""
    ):

        return """
        <script>
            alert("Lütfen Excel dosyası seçin.");
            window.location="/stoklar";
        </script>
        """

    try:

        # ==========================================
        # EXCELİ AÇ
        # ==========================================

        wb = openpyxl.load_workbook(
            dosya,
            data_only=True
        )

        ws = wb.active

        # ==========================================
        # EXCEL BAŞLIKLARI
        # ==========================================

        basliklar = []

        for hucre in ws[1]:

            if hucre.value is not None:

                basliklar.append(
                    str(
                        hucre.value
                    ).strip()
                )

            else:

                basliklar.append("")

        # ==========================================
        # GEREKLİ SÜTUNLAR
        # ==========================================

        gerekli_basliklar = [
            "Şube",
            "Ürün Adı",
            "Satış Miktarı"
        ]

        for baslik in gerekli_basliklar:

            if baslik not in basliklar:

                return f"""
                <script>

                    alert(
                        "Excel formatı hatalı.\\n\\n" +
                        "Eksik sütun: {baslik}\\n\\n" +
                        "Gerekli sütunlar:\\n" +
                        "Şube\\n" +
                        "Ürün Adı\\n" +
                        "Satış Miktarı"
                    );

                    window.location="/stoklar";

                </script>
                """

        # ==========================================
        # SÜTUN İNDEKSLERİ
        # ==========================================

        sube_index = basliklar.index(
            "Şube"
        )

        urun_index = basliklar.index(
            "Ürün Adı"
        )

        miktar_index = basliklar.index(
            "Satış Miktarı"
        )

        # ==========================================
        # SAYAÇLAR
        # ==========================================

        basarili = 0
        hatali = 0

        # ==========================================
        # SATIRLAR
        # ==========================================

        for satir in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            # ======================================
            # SATIR UZUNLUĞU
            # ======================================

            if len(satir) <= max(
                sube_index,
                urun_index,
                miktar_index
            ):

                hatali += 1
                continue

            sube_adi = satir[
                sube_index
            ]

            urun_adi = satir[
                urun_index
            ]

            miktar = satir[
                miktar_index
            ]

            # ======================================
            # BOŞ SATIR
            # ======================================

            if (
                sube_adi is None
                or urun_adi is None
                or miktar is None
            ):

                continue

            # ======================================
            # MİKTAR
            # ======================================

            try:

                miktar = float(
                    miktar
                )

            except (
                ValueError,
                TypeError
            ):

                hatali += 1
                continue

            if miktar <= 0:

                hatali += 1
                continue

            # ======================================
            # ŞUBE + ÜRÜN EŞLEŞTİR
            # ======================================

            eslesme = get_sube_urun_id(
                str(sube_adi).strip(),
                str(urun_adi).strip()
            )

            if not eslesme:

                hatali += 1
                continue

            sube_id = eslesme[0]
            urun_id = eslesme[1]

            # ======================================
            # ŞUBE STOKTAN SATIŞI DÜŞ
            # ======================================

            sube_stok_azalt(
                sube_id,
                urun_id,
                miktar
            )

            basarili += 1

        # ==========================================
        # SONUÇ
        # ==========================================

        return f"""
        <script>

            alert(
                "Satış aktarımı tamamlandı.\\n\\n" +
                "Başarılı kayıt: {basarili}\\n" +
                "Hatalı kayıt: {hatali}"
            );

            window.location="/stoklar";

        </script>
        """

    except Exception as e:

        return f"""
        <script>

            alert(
                "Excel aktarılırken hata oluştu:\\n\\n" +
                {json.dumps(
                    str(e),
                    ensure_ascii=False
                )}
            );

            window.location="/stoklar";

        </script>
        """


# ==========================================================
# EXCEL RAPOR
# ==========================================================

@app.route("/excel-rapor")
def excel_rapor():

    sonuc = admin_kontrol()

    if sonuc:
        return sonuc

    # ==========================================
    # EXCEL
    # ==========================================

    wb = Workbook()

    ws = wb.active

    ws.title = "Şube Siparişleri"

    urunler = get_tum_urunler()

    subeler = get_tum_subeler()

    # ==========================================
    # BAŞLIK
    # ==========================================

    ws["A1"] = "Ürün"

    ws["A1"].font = Font(
        bold=True
    )

    ws["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3"
    )

    ws["A1"].alignment = Alignment(
        horizontal="center"
    )

    sutun = 2

    for sube in subeler:

        hucre = ws.cell(
            row=1,
            column=sutun
        )

        hucre.value = sube[1]

        hucre.font = Font(
            bold=True
        )

        hucre.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3"
        )

        hucre.alignment = Alignment(
            horizontal="center"
        )

        sutun += 1

    # ==========================================
    # ÜRÜNLER
    # ==========================================

    satir = 2

    for urun in urunler:

        ws.cell(
            row=satir,
            column=1
        ).value = urun[1]

        sutun = 2

        for sube in subeler:

            miktar = get_siparis_miktari(
                sube[0],
                urun[0]
            )

            ws.cell(
                row=satir,
                column=sutun
            ).value = miktar

            sutun += 1

        satir += 1

    # ==========================================
    # FİLTRE
    # ==========================================

    ws.auto_filter.ref = ws.dimensions

    ws.freeze_panes = "B2"

    # ==========================================
    # SÜTUN GENİŞLİKLERİ
    # ==========================================

    ws.column_dimensions["A"].width = 35

    for column in range(
        2,
        ws.max_column + 1
    ):

        ws.column_dimensions[
            get_column_letter(column)
        ].width = 20

    # ==========================================
    # EXCEL DOSYASI
    # ==========================================

    dosya = BytesIO()

    wb.save(dosya)

    dosya.seek(0)

    return send_file(
        dosya,
        as_attachment=True,
        download_name="Siparis_Raporu.xlsx",
        mimetype=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


# ==========================================================
# UYGULAMA BAŞLANGIÇ KONTROLLERİ
# ==========================================================

print(
    "Login fonksiyonu =",
    login
)

print(
    "Endpoint =",
    app.view_functions["login"]
)

print("\n===== ROUTES =====")

for rule in app.url_map.iter_rules():

    print(
        rule,
        rule.methods
    )

print("==================\n")


# ==========================================================
# VERİTABANINI OLUŞTUR / GÜNCELLE
# ==========================================================

create_tables()


# ==========================================================
# UYGULAMAYI BAŞLAT
# ==========================================================

if __name__ == "__main__":

    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True
    )
    

